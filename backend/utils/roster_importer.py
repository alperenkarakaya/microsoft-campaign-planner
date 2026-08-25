"""
Roster importer: parse the curated creator spreadsheet and upsert into the DB.

Expects an .xlsx with a sheet named "List" and these columns (1-indexed):
  Col 1: Niche        — e.g. "Gaming", "Technology"
  Col 2: Name         — (often empty; display_name fallback)
  Col 3: Username     — @handle, e.g. "@BananaSlamJamma"
  Col 4: Profile Link — YouTube URL
  Col 5: Contact Info — business email
  Col 6: Location     — country/region string
  Col 7: Followers    — (usually empty; filled by enrichment)
  Col 8: Average views— (usually empty; filled by enrichment)
  Col 9: Talent agency— "YES" or "NO"
  Col 10: Approved    — (empty; deferred to CRM phase)

Idempotency: upserts by (platform, source_handle). Re-import updates
metadata but never duplicates a row.

Returns a summary dict with counts of created, updated, skipped, and
unresolved rows.
"""

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from sqlalchemy.orm import Session

from models import Influencer

logger = logging.getLogger(__name__)


@dataclass
class RosterRow:
    """One parsed row from the spreadsheet."""
    niche: Optional[str]
    name: Optional[str]
    handle: str              # normalized: lowercase, no leading @
    profile_link: Optional[str]
    email: Optional[str]
    location: Optional[str]
    talent_agency: Optional[bool]
    raw_handle: str          # as-is from the sheet, for audit


@dataclass
class ImportResult:
    """Summary of an import run."""
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: List[str] = field(default_factory=list)
    rows_parsed: int = 0


def _normalize_handle(raw: str) -> str:
    """Strip leading @, whitespace, and lowercase.

    >>> _normalize_handle("@BananaSlamJamma")
    'bananaslamjamma'
    >>> _normalize_handle("  @Hutts ")
    'hutts'
    """
    return raw.strip().lstrip("@").strip().lower()


def _parse_agency_flag(value) -> Optional[bool]:
    """Convert YES/NO/empty to bool/None.

    >>> _parse_agency_flag("YES")
    True
    >>> _parse_agency_flag("NO")
    False
    >>> _parse_agency_flag(None)
    """
    if value is None:
        return None
    s = str(value).strip().upper()
    if s == "YES":
        return True
    if s == "NO":
        return False
    return None


def _cell_str(value) -> Optional[str]:
    """Safely convert a cell value to a stripped string, or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_roster_xlsx(path: str, sheet_name: str = "List") -> List[RosterRow]:
    """Parse the roster spreadsheet and return a list of RosterRow objects.

    Skips rows where the Username (col 3) is empty or whitespace.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found; available: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows: List[RosterRow] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=10, values_only=True), start=2):
        # row is a tuple of cell values; columns are 0-indexed here
        raw_handle = row[2] if len(row) > 2 else None
        if not raw_handle or not str(raw_handle).strip():
            continue

        raw_handle_str = str(raw_handle).strip()
        handle = _normalize_handle(raw_handle_str)
        if not handle:
            continue

        rows.append(RosterRow(
            niche=_cell_str(row[0]) if len(row) > 0 else None,
            name=_cell_str(row[1]) if len(row) > 1 else None,
            handle=handle,
            profile_link=_cell_str(row[3]) if len(row) > 3 else None,
            email=_cell_str(row[4]) if len(row) > 4 else None,
            location=_cell_str(row[5]) if len(row) > 5 else None,
            talent_agency=_parse_agency_flag(row[8]) if len(row) > 8 else None,
            raw_handle=raw_handle_str,
        ))

    wb.close()
    logger.info("Parsed %d roster rows from %s", len(rows), path)
    return rows


def _extract_handle_from_url(url: str) -> Optional[str]:
    """Try to pull a @handle from a YouTube profile URL.

    >>> _extract_handle_from_url("https://www.youtube.com/@BananaSlamJamma/videos")
    'bananaslamjamma'
    """
    m = re.search(r"youtube\.com/@([^/?\s]+)", url)
    if m:
        return m.group(1).lower()
    return None


def upsert_roster(
    db: Session,
    rows: List[RosterRow],
    *,
    channel_resolver=None,
) -> ImportResult:
    """Upsert parsed roster rows into the influencers table.

    Args:
        db: SQLAlchemy session.
        rows: Parsed roster rows from parse_roster_xlsx.
        channel_resolver: Optional async callable(handle) -> channel_id or None.
            When None, channel resolution is skipped (handle stored as
            platform_id placeholder prefixed with "unresolved:").

    Idempotent: matches on (platform="youtube", source_handle=handle).
    """
    result = ImportResult(rows_parsed=len(rows))

    for row in rows:
        try:
            existing = (
                db.query(Influencer)
                .filter(
                    Influencer.platform == "youtube",
                    Influencer.source_handle == row.handle,
                )
                .first()
            )

            if existing:
                # Update mutable roster fields; don't clobber enrichment data.
                existing.business_email = row.email
                existing.talent_agency = row.talent_agency
                existing.category = row.niche or existing.category
                existing.country = row.location or existing.country
                if row.name and not existing.display_name:
                    existing.display_name = row.name
                result.updated += 1
            else:
                # platform_id is unique; use "unresolved:<handle>" until the
                # YouTube resolve step fills in the real channel ID.
                influencer = Influencer(
                    platform="youtube",
                    platform_id=f"unresolved:{row.handle}",
                    username=row.raw_handle,
                    display_name=row.name or row.raw_handle,
                    source_handle=row.handle,
                    business_email=row.email,
                    talent_agency=row.talent_agency,
                    category=row.niche,
                    country=row.location,
                    followers_count=0,
                    engagement_rate=0.0,
                )
                db.add(influencer)
                result.created += 1

        except Exception as e:
            logger.warning("Error upserting handle=%s: %s", row.handle, e)
            result.errors.append(f"{row.handle}: {e}")
            result.skipped += 1

    db.commit()
    logger.info(
        "Roster upsert done: %d created, %d updated, %d skipped, %d errors",
        result.created, result.updated, result.skipped, len(result.errors),
    )
    return result


async def resolve_handles(
    db: Session,
    *,
    dry_run: bool = False,
) -> Dict:
    """Resolve unresolved roster entries to YouTube channel IDs.

    Queries the YouTube API for each influencer whose platform_id starts with
    "unresolved:". On success, updates platform_id to the real channel ID and
    populates basic stats. On failure, the row is flagged but kept.

    Returns a summary dict: {resolved: int, failed: int, failures: [...]}.
    """
    import httpx

    from utils.youtube_api import YOUTUBE_API_KEY, YOUTUBE_API_BASE

    unresolved = (
        db.query(Influencer)
        .filter(
            Influencer.platform == "youtube",
            Influencer.platform_id.like("unresolved:%"),
        )
        .all()
    )

    if not unresolved:
        logger.info("No unresolved handles to resolve.")
        return {"resolved": 0, "failed": 0, "failures": []}

    if not YOUTUBE_API_KEY:
        logger.warning("YOUTUBE_API_KEY not set; skipping handle resolution.")
        return {
            "resolved": 0,
            "failed": len(unresolved),
            "failures": [f"{inf.source_handle}: YOUTUBE_API_KEY not set" for inf in unresolved],
        }

    resolved = 0
    failed = 0
    failures = []

    async with httpx.AsyncClient(timeout=20.0) as client:
        for inf in unresolved:
            handle = inf.source_handle
            try:
                # YouTube Data API v3: search for the @handle to find the channel.
                # forHandle is not a valid search param; use the channels endpoint
                # with forHandle parameter (available since 2023).
                resp = await client.get(
                    f"{YOUTUBE_API_BASE}/channels",
                    params={
                        "part": "snippet,statistics,contentDetails",
                        "forHandle": handle,
                        "key": YOUTUBE_API_KEY,
                    },
                )

                if resp.status_code != 200:
                    raise Exception(f"API returned {resp.status_code}")

                data = resp.json()
                items = data.get("items", [])

                if not items:
                    # Fallback: try searching by username
                    resp2 = await client.get(
                        f"{YOUTUBE_API_BASE}/search",
                        params={
                            "part": "snippet",
                            "q": f"@{handle}",
                            "type": "channel",
                            "maxResults": 1,
                            "key": YOUTUBE_API_KEY,
                        },
                    )
                    if resp2.status_code == 200:
                        search_items = resp2.json().get("items", [])
                        if search_items:
                            channel_id = search_items[0]["id"]["channelId"]
                            # Fetch full channel data
                            resp3 = await client.get(
                                f"{YOUTUBE_API_BASE}/channels",
                                params={
                                    "part": "snippet,statistics,contentDetails",
                                    "id": channel_id,
                                    "key": YOUTUBE_API_KEY,
                                },
                            )
                            if resp3.status_code == 200:
                                items = resp3.json().get("items", [])

                if not items:
                    failures.append(f"{handle}: channel not found")
                    failed += 1
                    continue

                channel = items[0]
                channel_id = channel["id"]
                stats = channel.get("statistics", {})
                snippet = channel.get("snippet", {})

                if not dry_run:
                    # Check if this channel_id already exists (dedup).
                    existing_by_channel = (
                        db.query(Influencer)
                        .filter(
                            Influencer.platform == "youtube",
                            Influencer.platform_id == channel_id,
                        )
                        .first()
                    )
                    if existing_by_channel and existing_by_channel.id != inf.id:
                        # Merge: keep the existing record, delete the unresolved dupe.
                        existing_by_channel.source_handle = inf.source_handle
                        existing_by_channel.business_email = inf.business_email or existing_by_channel.business_email
                        existing_by_channel.talent_agency = inf.talent_agency if inf.talent_agency is not None else existing_by_channel.talent_agency
                        db.delete(inf)
                        resolved += 1
                        continue

                    inf.platform_id = channel_id
                    inf.display_name = snippet.get("title", inf.display_name)
                    inf.username = snippet.get("customUrl", inf.username)
                    inf.bio = snippet.get("description", "")[:500] if snippet.get("description") else inf.bio
                    inf.avatar_url = snippet.get("thumbnails", {}).get("high", {}).get("url")
                    inf.followers_count = int(stats.get("subscriberCount", 0))
                    inf.country = snippet.get("country") or inf.country

                resolved += 1
                logger.debug("Resolved @%s -> %s", handle, channel_id)

            except Exception as e:
                logger.warning("Failed to resolve @%s: %s", handle, e)
                failures.append(f"{handle}: {e}")
                failed += 1

    if not dry_run:
        db.commit()

    logger.info("Handle resolution: %d resolved, %d failed", resolved, failed)
    return {"resolved": resolved, "failed": failed, "failures": failures}
